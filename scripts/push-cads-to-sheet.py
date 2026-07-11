"""
Day du lieu tu file Excel xuat CADS (CADSDocuments*.xlsx) thang vao Google Sheet
"don hang nhap" (sheet B2B) ma app dieu-xe dang doc.

CHUAN BI (lam 1 lan):
  1. Vao https://console.cloud.google.com/ -> tao project (hoac dung project co san).
  2. Bat API "Google Sheets API" cho project do.
  3. Vao IAM & Admin > Service Accounts -> Create Service Account -> dat ten gi cung duoc.
  4. Vao tab "Keys" cua service account -> Add Key -> Create new key -> chon JSON -> tai file ve.
  5. Mo file JSON vua tai, copy gia tri "client_email" (dang xxx@xxx.iam.gserviceaccount.com).
  6. Mo Google Sheet don hang -> Chia se (Share) -> dan email do vao, chon quyen "Nguoi chinh sua" (Editor).
  7. Dat file JSON da tai vao cung thu muc `scripts/` nay, doi ten thanh `service-account.json`
     (file nay KHONG duoc commit len git -- da them vao .gitignore).
  8. Cai thu vien can thiet:  pip install gspread google-auth

CHAY:
  python scripts/push-cads-to-sheet.py "duong/dan/CADSDocuments26155914.xlsx"

  Mac dinh se APPEND (them dong moi) vao cuoi sheet "Sheet1" (gid=0) cua:
  https://docs.google.com/spreadsheets/d/1cpSsk_kUbJ3Yy_g-UVNyvWrv0sHtdKRdxunpPwNstV4

  Them flag --dry-run de chi in ra cac dong se duoc ghi, KHONG ghi thiat vao sheet
  (dung de kiem tra truoc khi chay thiat):
  python scripts/push-cads-to-sheet.py "file.xlsx" --dry-run

QUY TAC MAP COT (da thong nhat voi nguoi dung ngay 2026-06-26):
  - File CADS chi co So luong TONG theo hoa don, khong tach loai hang (A3/A4/A5/Vo/Giay ve sinh)
    -> cac cot san pham do de TRONG, nguoi dung tu dien tay sau.
  - So luong + Dien giai duoc gop va ghi vao cot GHI CHU de doi chieu.
  - NGAY CAN GIAO / NGAY GIAO: khong co trong file CADS -> de trong.
  - DIA CHI: neu file nguon khong co du lieu (cot rong), van day len voi DIA CHI trong --
    nguoi dung tu dien tay sau khi doi chieu ho so goc.

NHAN DIEN COT: script tu do cot nguon theo ten da chuan hoa (bo dau, viet hoa -- giong
ham norm() trong lib/sheets.js), nen van hoat dong neu cac lan xuat CADS sau co ten cot
hoi khac (vi du "So HD" / "SoHD" / "So Hop Dong" deu nhan ra cung 1 field "so_hd").
Moi lan chay se in ra bang "Nhan dien cot nguon" de kiem tra truoc khi tin tuong ket qua.
"""

import re
import unicodedata
import argparse
from datetime import datetime

import openpyxl

SHEET_ID = "1cpSsk_kUbJ3Yy_g-UVNyvWrv0sHtdKRdxunpPwNstV4"
SHEET_TAB_GID = 0  # tab "Sheet1" / B2B dang dung cho dieu-xe
SERVICE_ACCOUNT_FILE = "scripts/service-account.json"

# Thu tu cot dung nhu header hien tai cua Google Sheet (khong tu doi thu tu)
TARGET_HEADERS = [
    "SỐ PHIÊU", "NGÀY LÊN ĐƠN", "NGÀY CẦN GIAO", "KHÁCH HÀNG", "ĐỊA CHỈ",
    "SĐT NGƯỜI NHẬN", "NGÀY GIAO", "SẢN PHẨM A3", "SẢN PHẨM A4", "SẢN PHẨM A5",
    "VỞ", "GIẤY VỆ SINH", "LOẠI XUẤT", "GHI CHÚ", "LÁI XE", "GIAO NHẬN",
]

# Nhan dien cot nguon: moi field nguon co the xuat hien voi ten hoi khac nhau
# giua cac lan xuat CADS (doi merchant, doi version ERP...) -> match "khop dung"
# truoc, "chua tu khoa" sau, de van nhan dien duoc khi ten cot doi nhe.
FIELD_CANDIDATES = {
    "so_hd":      {"exact": ["SO HD"], "contains": ["SO HD"], "exclude_contains": ["CN", "KY HIEU"]},
    "ngay_hd":    {"exact": ["NGAY HD"], "contains": ["NGAY HD", "NGAY LEN DON", "NGAY CHUNG TU"], "exclude_contains": ["CN"]},
    "ten_kh":     {"exact": ["TEN KHACH HANG"], "contains": ["TEN KHACH HANG", "KHACH HANG"], "exclude_contains": ["MA"]},
    "dia_chi":    {"exact": ["DIA CHI GIAO HANG"], "contains": ["DIA CHI GIAO", "DIA CHI"], "exclude_contains": []},
    "nguoi_mua":  {"exact": ["NGUOI MUA"], "contains": ["NGUOI MUA", "NGUOI NHAN", "NGUOI LIEN HE"], "exclude_contains": []},
    "so_luong":   {"exact": ["SO LUONG"], "contains": ["SO LUONG"], "exclude_contains": []},
    "dien_giai":  {"exact": ["DIEN GIAI"], "contains": ["DIEN GIAI"], "exclude_contains": []},
}


def norm(h):
    """Chuan hoa header: bo dau, viet hoa, gop khoang trang -- cung logic norm() trong lib/sheets.js."""
    s = str(h or "").upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("Đ", "D")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def detect_columns(headers):
    """Tra ve (dict field -> index cot, list field khong nhan dien duoc)."""
    normed = [norm(h) for h in headers]
    found = {}
    missing = []
    for field, rule in FIELD_CANDIDATES.items():
        idx = None
        for cand in rule["exact"]:
            for i, h in enumerate(normed):
                if h == cand:
                    idx = i
                    break
            if idx is not None:
                break
        if idx is None:
            for cand in rule["contains"]:
                for i, h in enumerate(normed):
                    if cand in h and not any(ex in h for ex in rule["exclude_contains"]):
                        idx = i
                        break
                if idx is not None:
                    break
        if idx is None:
            missing.append(field)
        else:
            found[field] = idx
    return found, missing


def fmt_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def clean_dien_giai(v):
    if not v:
        return ""
    lines = [l.strip() for l in str(v).split("\n") if l.strip()]
    return lines[0] if lines else ""


def read_cads_rows(xlsx_path, verbose=True):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet"] if "Sheet" in wb.sheetnames else wb.worksheets[0]
    headers = [c.value for c in ws[1]]

    col, missing = detect_columns(headers)

    if verbose:
        print("Nhan dien cot nguon:")
        for field in FIELD_CANDIDATES:
            if field in col:
                print(f"  {field:12s} <- cot '{headers[col[field]]}' (vi tri {col[field] + 1})")
            else:
                print(f"  {field:12s} <- KHONG TIM THAY, se de trong")

    if "so_hd" in missing:
        raise SystemExit(
            "Khong tim thay cot 'So HD' (ma don) trong file -- khong the tach dong du lieu. Dung lai."
        )

    def get(row, field):
        i = col.get(field)
        return row[i] if i is not None and i < len(row) else None

    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        so_hd = get(row, "so_hd")
        if not so_hd:
            continue
        ten_kh    = get(row, "ten_kh") or ""
        dia_chi   = get(row, "dia_chi") or ""
        nguoi_mua = get(row, "nguoi_mua") or ""
        so_luong  = get(row, "so_luong") or 0
        dien_giai = clean_dien_giai(get(row, "dien_giai"))
        ngay_hd   = fmt_date(get(row, "ngay_hd"))

        sdt = re.sub(r"\s+", " ", str(nguoi_mua).replace("\t", " - ")).strip()
        ghi_chu = f"SL: {so_luong}" + (f" | {dien_giai}" if dien_giai else "")

        sheet_row = {
            "SỐ PHIÊU": str(so_hd).strip(),
            "NGÀY LÊN ĐƠN": ngay_hd,
            "NGÀY CẦN GIAO": "",
            "KHÁCH HÀNG": str(ten_kh).strip(),
            "ĐỊA CHỈ": str(dia_chi).strip(),
            "SĐT NGƯỜI NHẬN": sdt,
            "NGÀY GIAO": "",
            "SẢN PHẨM A3": "",
            "SẢN PHẨM A4": "",
            "SẢN PHẨM A5": "",
            "VỞ": "",
            "GIẤY VỆ SINH": "",
            "LOẠI XUẤT": "",
            "GHI CHÚ": ghi_chu,
            "LÁI XE": "",
            "GIAO NHẬN": "",
        }
        rows_out.append([sheet_row[h] for h in TARGET_HEADERS])
    return rows_out


def push_to_sheet(rows):
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)
    ws = next((w for w in sh.worksheets() if w.id == SHEET_TAB_GID), sh.sheet1)
    ws.append_rows(rows, value_input_option="USER_ENTERED")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    rows = read_cads_rows(args.xlsx_path)
    print(f"\nDoc duoc {len(rows)} dong tu {args.xlsx_path}")
    for r in rows:
        print(r)

    if args.dry_run:
        print("\n--dry-run: chua ghi gi vao Google Sheet.")
        return

    push_to_sheet(rows)
    print(f"\nDa them {len(rows)} dong vao Google Sheet.")


if __name__ == "__main__":
    main()
